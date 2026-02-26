from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest

from agrobr.exceptions import ParseError
from agrobr.mapbiomas.parser import (
    PARSER_VERSION,
    parse_cobertura_xlsx,
    parse_transicao_xlsx,
)

GOLDEN_DIR = Path(__file__).parent.parent / "golden_data" / "mapbiomas"


def _golden_xlsx() -> bytes:
    return GOLDEN_DIR.joinpath("biome_state_sample.xlsx").read_bytes()


def _expected() -> dict:
    return json.loads(GOLDEN_DIR.joinpath("expected.json").read_text(encoding="utf-8"))


class TestParserVersion:
    def test_version_is_int(self):
        assert isinstance(PARSER_VERSION, int)
        assert PARSER_VERSION >= 1


class TestParseCoberturaXlsx:
    def test_valid_xlsx(self):
        df = parse_cobertura_xlsx(_golden_xlsx())

        assert len(df) >= 20
        assert "bioma" in df.columns
        assert "estado" in df.columns
        assert "classe_id" in df.columns
        assert "classe" in df.columns
        assert "ano" in df.columns
        assert "area_ha" in df.columns

    def test_golden_data_columns(self):
        expected = _expected()["cobertura"]
        df = parse_cobertura_xlsx(_golden_xlsx())

        for col in expected["columns"]:
            assert col in df.columns, f"Missing column: {col}"

    def test_golden_data_biomas(self):
        expected = _expected()["cobertura"]
        df = parse_cobertura_xlsx(_golden_xlsx())

        biomas = sorted(df["bioma"].unique().tolist())
        for b in expected["biomas_expected"]:
            assert b in biomas, f"Missing bioma: {b}"

    def test_golden_data_estados(self):
        expected = _expected()["cobertura"]
        df = parse_cobertura_xlsx(_golden_xlsx())

        estados = sorted(df["estado"].unique().tolist())
        for e in expected["estados_expected"]:
            assert e in estados, f"Missing estado: {e}"

    def test_golden_data_anos(self):
        expected = _expected()["cobertura"]
        df = parse_cobertura_xlsx(_golden_xlsx())

        anos = sorted(df["ano"].dropna().unique().tolist())
        for a in expected["anos_expected"]:
            assert a in anos, f"Missing ano: {a}"

    def test_area_non_negative(self):
        df = parse_cobertura_xlsx(_golden_xlsx())
        assert (df["area_ha"] >= 0).all()

    def test_ano_is_numeric(self):
        df = parse_cobertura_xlsx(_golden_xlsx())
        assert pd.api.types.is_integer_dtype(df["ano"])

    def test_classe_id_is_numeric(self):
        df = parse_cobertura_xlsx(_golden_xlsx())
        assert pd.api.types.is_integer_dtype(df["classe_id"])

    def test_invalid_xlsx_raises(self):
        with pytest.raises(ParseError):
            parse_cobertura_xlsx(b"invalid data")

    def test_empty_xlsx_raises(self):
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COVERAGE_10"
        ws.append(["biome", "state", "class", "class_level_0"])
        buf = BytesIO()
        wb.save(buf)
        with pytest.raises(ParseError, match="COVERAGE vazia"):
            parse_cobertura_xlsx(buf.getvalue())

    def test_missing_columns_raises(self):
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COVERAGE_10"
        ws.append(["id", "nome", "valor"])
        ws.append([1, "teste", 100])
        buf = BytesIO()
        wb.save(buf)
        with pytest.raises(ParseError, match="Colunas obrigatorias ausentes"):
            parse_cobertura_xlsx(buf.getvalue())


class TestParseCoberturaMunicipal:
    def _make_municipal_xlsx(self) -> bytes:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COVERAGE_10"
        ws.append(["biome", "state", "municipality", "class", "class_level_0", 2020, 2021])
        ws.append(["Amazônia", "Pará", "Belém", 3, "Natural", 100.0, 110.0])
        ws.append(["Amazônia", "Pará", "Marabá", 15, "Antropic", 200.0, 190.0])
        ws.append(["Cerrado", "Goiás", "Goiânia", 3, "Natural", 150.0, 160.0])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_municipality_column_present(self):
        df = parse_cobertura_xlsx(self._make_municipal_xlsx())
        assert "municipio" in df.columns

    def test_municipality_values(self):
        df = parse_cobertura_xlsx(self._make_municipal_xlsx())
        municipios = df["municipio"].unique().tolist()
        assert "Belém" in municipios
        assert "Marabá" in municipios
        assert "Goiânia" in municipios

    def test_municipality_output_columns(self):
        df = parse_cobertura_xlsx(self._make_municipal_xlsx())
        expected = [
            "bioma",
            "estado",
            "municipio",
            "classe_id",
            "classe",
            "nivel_0",
            "ano",
            "area_ha",
        ]
        assert list(df.columns) == expected

    def test_municipality_records_count(self):
        df = parse_cobertura_xlsx(self._make_municipal_xlsx())
        assert len(df) == 6  # 3 rows x 2 years

    def test_state_level_no_municipality(self):
        df = parse_cobertura_xlsx(_golden_xlsx())
        assert "municipio" not in df.columns


class TestParseTransicaoXlsx:
    def test_valid_xlsx(self):
        df = parse_transicao_xlsx(_golden_xlsx())

        assert len(df) >= 20
        assert "bioma" in df.columns
        assert "estado" in df.columns
        assert "classe_de_id" in df.columns
        assert "classe_de" in df.columns
        assert "classe_para_id" in df.columns
        assert "classe_para" in df.columns
        assert "periodo" in df.columns
        assert "area_ha" in df.columns

    def test_golden_data_columns(self):
        expected = _expected()["transicao"]
        df = parse_transicao_xlsx(_golden_xlsx())

        for col in expected["columns"]:
            assert col in df.columns, f"Missing column: {col}"

    def test_golden_data_biomas(self):
        expected = _expected()["transicao"]
        df = parse_transicao_xlsx(_golden_xlsx())

        biomas = sorted(df["bioma"].unique().tolist())
        for b in expected["biomas_expected"]:
            assert b in biomas, f"Missing bioma: {b}"

    def test_golden_data_estados(self):
        expected = _expected()["transicao"]
        df = parse_transicao_xlsx(_golden_xlsx())

        estados = sorted(df["estado"].unique().tolist())
        for e in expected["estados_expected"]:
            assert e in estados, f"Missing estado: {e}"

    def test_golden_data_periodos(self):
        expected = _expected()["transicao"]
        df = parse_transicao_xlsx(_golden_xlsx())

        periodos = sorted(df["periodo"].unique().tolist())
        for p in expected["periodos_expected"]:
            assert p in periodos, f"Missing periodo: {p}"

    def test_periodo_format(self):
        df = parse_transicao_xlsx(_golden_xlsx())
        for p in df["periodo"].unique():
            parts = p.split("-")
            assert len(parts) == 2
            assert parts[0].isdigit()
            assert parts[1].isdigit()

    def test_area_non_negative(self):
        df = parse_transicao_xlsx(_golden_xlsx())
        assert (df["area_ha"] >= 0).all()

    def test_classe_ids_are_numeric(self):
        df = parse_transicao_xlsx(_golden_xlsx())
        assert pd.api.types.is_integer_dtype(df["classe_de_id"])
        assert pd.api.types.is_integer_dtype(df["classe_para_id"])

    def test_invalid_xlsx_raises(self):
        with pytest.raises(ParseError):
            parse_transicao_xlsx(b"invalid data")

    def test_missing_columns_raises(self):
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TRANSITION_10"
        ws.append(["id", "nome", "valor"])
        ws.append([1, "teste", 100])
        buf = BytesIO()
        wb.save(buf)
        with pytest.raises(ParseError, match="Colunas obrigatorias ausentes"):
            parse_transicao_xlsx(buf.getvalue())
