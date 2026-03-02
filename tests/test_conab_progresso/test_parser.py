from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from agrobr.conab.progresso.parser import PARSER_VERSION, parse_progresso_xlsx
from agrobr.exceptions import ParseError

GOLDEN_DIR = Path(__file__).resolve().parent.parent / "golden_data" / "conab_progresso"


@pytest.fixture()
def golden_xlsx() -> bytes:
    return (GOLDEN_DIR / "progresso_sample" / "response.xlsx").read_bytes()


@pytest.fixture()
def expected() -> dict:
    return json.loads(
        (GOLDEN_DIR / "progresso_sample" / "expected.json").read_text(encoding="utf-8")
    )


class TestParserVersion:
    def test_parser_version(self) -> None:
        assert PARSER_VERSION == 1


class TestParseGolden:
    def test_columns(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert df.columns.tolist() == expected["columns"]

    def test_record_count(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert len(df) == expected["total_records"]

    def test_culturas(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert sorted(df["cultura"].unique().tolist()) == sorted(expected["culturas"])

    def test_operacoes(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert sorted(df["operacao"].unique().tolist()) == sorted(expected["operacoes"])

    def test_estados(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert sorted(df["estado"].unique().tolist()) == sorted(expected["estados"])

    def test_semana(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert df["semana_atual"].unique().tolist() == [expected["semana"]]

    def test_safra(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert df["safra"].unique().tolist() == [expected["safra"]]

    def test_soja_semeadura_count(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        soja_sem = df[(df["cultura"] == "Soja") & (df["operacao"] == "Semeadura")]
        assert len(soja_sem) == expected["soja_semeadura_records"]

    def test_soja_colheita_count(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        soja_col = df[(df["cultura"] == "Soja") & (df["operacao"] == "Colheita")]
        assert len(soja_col) == expected["soja_colheita_records"]

    def test_milho2_semeadura_count(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        milho2 = df[df["cultura"].str.contains("Milho 2")]
        assert len(milho2) == expected["milho2_semeadura_records"]

    def test_mt_soja_colheita(self, golden_xlsx: bytes, expected: dict) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        row = df[
            (df["cultura"] == "Soja") & (df["estado"] == "MT") & (df["operacao"] == "Colheita")
        ]
        assert len(row) == 1
        assert row.iloc[0]["pct_semana_atual"] == pytest.approx(
            expected["mt_soja_colheita_pct_atual"]
        )

    def test_values_between_0_and_1(self, golden_xlsx: bytes) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        for col in [
            "pct_ano_anterior",
            "pct_semana_anterior",
            "pct_semana_atual",
            "pct_media_5_anos",
        ]:
            valid = df[col].dropna()
            assert (valid >= 0).all(), f"{col} has negative values"
            assert (valid <= 1.01).all(), f"{col} has values > 1"


class TestParseEdgeCases:
    def test_empty_xlsx_raises(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progresso de safra"
        import io

        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ParseError, match="Sheet vazia"):
            parse_progresso_xlsx(buf.getvalue())

    def test_no_records_raises(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progresso de safra"
        ws["B1"] = "Header only"
        ws["B2"] = "No data here"
        import io

        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ParseError, match="Nenhum registro"):
            parse_progresso_xlsx(buf.getvalue())

    def test_invalid_bytes_raises(self) -> None:
        with pytest.raises(ParseError, match="Erro ao abrir Excel"):
            parse_progresso_xlsx(b"not an xlsx file")

    def test_returns_dataframe(self, golden_xlsx: bytes) -> None:
        result = parse_progresso_xlsx(golden_xlsx)
        assert isinstance(result, pd.DataFrame)

    def test_aggregate_rows_as_br(self, golden_xlsx: bytes) -> None:
        df = parse_progresso_xlsx(golden_xlsx)
        assert "BR" in df["estado"].values
        for estado in df["estado"]:
            assert "estados" not in str(estado).lower()
